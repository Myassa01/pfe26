"""StructuredQueryEngine — interroge les fichiers Excel via DuckDB en mémoire."""
from __future__ import annotations

import logging
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class StructuredQueryEngine:
    EXCEL_EXTS = {".xlsx", ".xls"}

    def __init__(self, docs_dir: str):
        import duckdb
        self._duckdb = duckdb
        self.conn = duckdb.connect(":memory:")
        self.tables: Dict[str, Dict[str, Any]] = {}
        self.last_warnings: List[str] = []
        self.docs_dir = docs_dir
        self._load_all(docs_dir)

    # ── Chargement ────────────────────────────────────────────────────────

    def _load_all(self, docs_dir: str) -> None:
        if not Path(docs_dir).exists():
            logger.warning("StructuredQueryEngine: docs_dir introuvable: %s", docs_dir)
            return
        loaded = 0
        for file in sorted(Path(docs_dir).rglob("*")):
            if not file.is_file() or file.suffix.lower() not in self.EXCEL_EXTS:
                continue
            try:
                self._load_excel(file)
                loaded += 1
            except Exception as e:
                logger.warning("  ⚠ Skip %s: %s", file.name, e)
        logger.info("StructuredQueryEngine: %d table(s) chargée(s) en DuckDB (%s)",
                    loaded, ", ".join(self.tables.keys()) or "—")

    def _load_excel(self, path: Path) -> None:
        from ..ingestion.loader import _detect_header_row
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            ws = wb.active or wb.worksheets[0]

            header_row = _detect_header_row(ws)
            headers: List[str] = []
            for cell in ws[header_row]:
                headers.append(str(cell.value).strip() if cell.value else "")

            keep_idx = [i for i, h in enumerate(headers) if h]
            sql_headers = [self._sql_ident(headers[i]) for i in keep_idx]
            if not sql_headers:
                logger.warning("  ⚠ %s: aucune colonne valide", path.name)
                return

            seen: Dict[str, int] = {}
            unique_headers: List[str] = []
            for h in sql_headers:
                if h in seen:
                    seen[h] += 1
                    unique_headers.append(f"{h}_{seen[h]}")
                else:
                    seen[h] = 0
                    unique_headers.append(h)

            rows: List[List[Optional[str]]] = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                values = [row[i] if i < len(row) else None for i in keep_idx]
                values = [None if v is None else str(v).strip() for v in values]
                if any(v for v in values):
                    rows.append(values)
        finally:
            wb.close()

        if not rows:
            logger.warning("  ⚠ %s: aucune donnée", path.name)
            return

        table_name = self._normalize_stem(path.name)
        sql_table = self._sql_ident(table_name)

        cols_def = ", ".join(f'"{c}" VARCHAR' for c in unique_headers)
        self.conn.execute(f'CREATE OR REPLACE TABLE "{sql_table}" ({cols_def})')

        placeholders = ", ".join(["?"] * len(unique_headers))
        self.conn.executemany(
            f'INSERT INTO "{sql_table}" VALUES ({placeholders})',
            rows,
        )

        technical_cols = self._detect_technical_columns(sql_table, unique_headers, len(rows))

        column_samples: Dict[str, List[str]] = {}
        for col in unique_headers:
            try:
                rs = self.conn.execute(
                    f'SELECT DISTINCT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' LIMIT 5'
                ).fetchall()
                column_samples[col] = [str(r[0]).strip() for r in rs if r[0]]
            except Exception:
                column_samples[col] = []

        self.tables[table_name] = {
            "filename": path.name,
            "columns": unique_headers,
            "technical_columns": technical_cols,
            "user_columns": [c for c in unique_headers if c not in technical_cols],
            "row_count": len(rows),
            "sql_table": sql_table,
            "column_samples": column_samples,
        }
        if technical_cols:
            logger.info("  ✓ %s → table %s (%d lignes, %d colonnes ; techniques: %s)",
                        path.name, sql_table, len(rows), len(unique_headers),
                        ", ".join(sorted(technical_cols)))
        else:
            logger.info("  ✓ %s → table %s (%d lignes, %d colonnes)",
                        path.name, sql_table, len(rows), len(unique_headers))

    # Pattern to detect technical/code column names — covers common DB conventions
    # without any business-domain vocabulary.
    _TECH_NAME_RE = re.compile(
        r"^(ID|CODE|CD|NUM|REF|NO|N|NBR|CIN|NIN|MATRICULE|IMMATRICULATION)$"
        r"|^(ID|CD|NUM|REF|NO|NBR)_"
        r"|_(ID|CODE|CD|NUM|REF|NO)$",
        re.IGNORECASE,
    )

    # Colonnes dont le nom suggère un attribut métier numérique — ne jamais
    # les marquer comme techniques même si leurs valeurs sont numériques.
    # Couverture volontairement large et multilingue : SAL/WAGE/PAY, DATE/YEAR,
    # TEL/PHONE/MOBILE, MONT/AMOUNT, etc.
    _BUSINESS_NUM_RE = re.compile(
        r"SAL|WAGE|PAY|REMUN"
        r"|DATE|YEAR|NAIS|BIRTH"
        r"|AGE|DUR|ANCIEN|EXP"
        r"|TEL|PHONE|MOBILE|GSM"
        r"|MONT|AMT|AMOUNT|PRIX|PRICE"
        r"|MAIL|EMAIL",
        re.IGNORECASE,
    )

    # Mots vides grammaticaux pour find_column_for_question — aucun vocabulaire métier.
    _MATCH_STOP = frozenset({
        "qui", "quel", "quelle", "quels", "quelles", "est", "sont",
        "le", "la", "les", "un", "une", "des", "du",
        "de", "d", "l", "j", "m", "n", "s", "t", "y", "c",
        "ce", "cet", "cette", "ces",
        "dans", "par", "pour", "sur", "avec", "sans", "en", "au", "aux",
        "et", "ou", "mais", "donc", "car", "si",
        "combien", "comment", "quand", "pourquoi", "que", "qu",
        "the", "is", "are", "of", "to", "in", "for", "with",
        "what", "who", "how", "when", "where", "which",
        "a", "an", "and", "or",
    })

    def _detect_technical_columns(self, sql_table: str, columns: List[str], row_count: int) -> set:
        if row_count == 0:
            return set()
        sample_size = min(row_count, 200)
        technical: set = set()
        for col in columns:
            if self._TECH_NAME_RE.search(col):
                technical.add(col)
                continue
            # Attribut métier numérique connu → jamais technique
            if self._BUSINESS_NUM_RE.search(col):
                continue
            try:
                rows = self.conn.execute(
                    f'SELECT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' '
                    f'LIMIT {sample_size}'
                ).fetchall()
            except Exception:
                continue
            if not rows:
                continue
            vals = [str(r[0]).strip() for r in rows]
            n = len(vals)
            numeric_vals: List[int] = []
            for v in vals:
                clean = re.sub(r"\s", "", v).replace(",", "").replace(".", "")
                if clean.lstrip("-").isdigit() and clean:
                    try:
                        numeric_vals.append(abs(int(clean.lstrip("-"))))
                    except ValueError:
                        pass
            numeric_ratio = len(numeric_vals) / n
            short_alphanum = sum(
                1 for v in vals
                if len(v) <= 6 and v.replace("_", "").replace("-", "").isalnum()
            )
            if numeric_ratio > 0.7:
                # Colonne année (ex : DATE_NAISS) : valeurs 4 chiffres [1900-2035] → garder
                if numeric_vals and all(1900 <= v <= 2035 for v in numeric_vals):
                    continue
                # Identifiant séquentiel : max value ≤ seuil → technique
                # Salaires (50 000+), téléphones (055…) dépassent ce seuil → gardés
                id_threshold = max(row_count * 10, 9999)
                if max(numeric_vals, default=0) <= id_threshold:
                    technical.add(col)
            elif short_alphanum / n > 0.7 and numeric_ratio > 0.3:
                technical.add(col)
        return technical

    # ── Helpers normalisation ─────────────────────────────────────────────

    @staticmethod
    def _normalize_stem(fname: str) -> str:
        stem = fname.rsplit(".", 1)[0] if "." in fname else fname
        stem = stem.upper().strip()
        stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
        stem = re.sub(r"\s*_\d+\s*$", "", stem)
        return stem.strip()

    @staticmethod
    def _sql_ident(name: str) -> str:
        nfd = unicodedata.normalize("NFD", name)
        ascii_name = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
        cleaned = re.sub(r"[^A-Za-z0-9_]+", "_", ascii_name).strip("_").upper()
        if not cleaned:
            cleaned = "COL"
        if cleaned[0].isdigit():
            cleaned = "C_" + cleaned
        return cleaned

    @staticmethod
    def _fold(text: str) -> str:
        nfd = unicodedata.normalize("NFD", str(text))
        ascii_text = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
        # Strip straight (U+0027) and curly (U+2019) apostrophes for agnostic matching.
        # "H.RMEL" (no apostrophe) now matches "H.R'MEL" (with apostrophe).
        return ascii_text.lower().strip().replace("'", "").replace("’", "")

    @staticmethod
    def _like_col(col: str) -> str:
        """SQL expression for LIKE: accent + case + apostrophe normalization on a column."""
        return (
            f"replace(replace(strip_accents(LOWER(CAST(\"{col}\" AS VARCHAR))), "
            f"chr(39), ''), chr(8217), '')"
        )

    # Same normalization applied to the ? parameter side.
    _LIKE_PARAM = "replace(replace(strip_accents(LOWER(?)), chr(39), ''), chr(8217), '')"

    def _resolve_column(self, table: str, column: Optional[str]) -> Optional[str]:
        if not column:
            return None
        cols = self.tables[table]["columns"]
        target = self._fold(column)
        target_sql = self._sql_ident(column)
        for c in cols:
            if self._fold(c) == target or c == target_sql:
                return c
        return None

    # ── API publique ──────────────────────────────────────────────────────

    def has_table(self, name: str) -> bool:
        return self._normalize_stem(name) in self.tables \
            or name in self.tables

    def schema(self) -> Dict[str, Dict[str, Any]]:
        return {
            name: {
                "columns": info["columns"],
                "filename": info["filename"],
                "row_count": info["row_count"],
                "is_doc": False,
            }
            for name, info in self.tables.items()
        }

    def samples(self, table: str, max_per_col: int = 3, max_len: int = 40) -> Dict[str, List[str]]:
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return {}
        sql_table = self.tables[table]["sql_table"]
        out: Dict[str, List[str]] = {}
        for col in self.tables[table]["columns"]:
            try:
                rows = self.conn.execute(
                    f'SELECT DISTINCT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL LIMIT {max_per_col}'
                ).fetchall()
                vals = [str(r[0])[:max_len] for r in rows if r[0]]
                if vals:
                    out[col] = vals
            except Exception:
                continue
        return out

    def list_values(
        self,
        table: str,
        column: Optional[str] = None,
        filters: Optional[Dict[str, str]] = None,
        distinct: bool = True,
        hide_technical: bool = True,
    ) -> List[Dict[str, Any]]:
        self.last_warnings = []
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return []
        sql_table = self.tables[table]["sql_table"]
        meta_filename = self.tables[table]["filename"]
        all_cols = self.tables[table]["columns"]

        resolved_filters: List[tuple] = []
        dropped_filters: List[str] = []
        for fkey, fval in (filters or {}).items():
            sql_col_filter = self._resolve_column(table, fkey)
            if sql_col_filter and fval:
                resolved_filters.append((sql_col_filter, str(fval)))
            else:
                dropped_filters.append(f"{fkey}={fval!r}")
        if dropped_filters:
            msg = (f"Filtre(s) ignoré(s) : {', '.join(dropped_filters)} "
                   f"— colonne(s) absente(s) de la table {table} "
                   f"(colonnes disponibles : {', '.join(all_cols)}).")
            logger.warning("  ⚠ %s", msg)
            self.last_warnings.append(msg)

        sql_col = self._resolve_column(table, column) if column else None

        if sql_col:
            select_clause = f'"{sql_col}"'
        else:
            if hide_technical:
                cols = self.tables[table].get("user_columns") or self.tables[table]["columns"]
            else:
                cols = self.tables[table]["columns"]
            select_clause = ", ".join(f'"{c}"' for c in cols)

        distinct_kw = "DISTINCT " if distinct else ""

        where_parts = []
        params: List[str] = []
        if sql_col:
            where_parts.append(f'"{sql_col}" IS NOT NULL')
            where_parts.append(f'TRIM("{sql_col}") <> \'\'')
        for col, val in resolved_filters:
            where_parts.append(
                f"{self._like_col(col)} LIKE {self._LIKE_PARAM}"
            )
            params.append(f"%{val}%")

        where_clause = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        sql = f'SELECT {distinct_kw}{select_clause} FROM "{sql_table}"{where_clause}'

        try:
            cursor = self.conn.execute(sql, params)
            rows = cursor.fetchall()
            col_names = [d[0] for d in cursor.description]
        except Exception as e:
            logger.warning("StructuredQueryEngine: requête SQL échouée (%s) — fallback lower", e)
            return self._list_values_fallback(table, column, filters, distinct)

        if not rows and resolved_filters:
            tokenized_where = []
            tokenized_params: List[str] = []
            if sql_col:
                tokenized_where.append(f'"{sql_col}" IS NOT NULL')
                tokenized_where.append(f'TRIM("{sql_col}") <> \'\'')
            had_tokens = False
            for col, val in resolved_filters:
                # Split sur espaces ET points (gère les abréviations H.R'MEL → R'MEL)
                tokens = [t for t in re.split(r"[\s.]+", val.strip()) if len(t) >= 3]
                tokens = [t[:8] if len(t) > 8 else t for t in tokens]
                for t in tokens:
                    tokenized_where.append(
                        f"{self._like_col(col)} LIKE {self._LIKE_PARAM}"
                    )
                    tokenized_params.append(f"%{t}%")
                    had_tokens = True
            if had_tokens:
                fallback_sql = (
                    f'SELECT {distinct_kw}{select_clause} FROM "{sql_table}" '
                    f'WHERE ' + " AND ".join(tokenized_where)
                )
                try:
                    cursor = self.conn.execute(fallback_sql, tokenized_params)
                    rows = cursor.fetchall()
                    col_names = [d[0] for d in cursor.description]
                    if rows:
                        msg = (f"Aucun résultat pour le filtre exact, "
                               f"{len(rows)} résultat(s) en match approximatif "
                               f"(tokens: {tokenized_params}).")
                        logger.info("  ⟹ %s", msg)
                        self.last_warnings.append(msg)
                except Exception as e:
                    logger.warning("Fallback tokenisé échoué: %s", e)

        results: List[Dict[str, Any]] = []
        for row in rows:
            if sql_col:
                val = row[0]
                if val is None or not str(val).strip():
                    continue
                results.append({
                    "content": str(val).strip(),
                    "metadata": {"filename": meta_filename, "table": table, "column": sql_col},
                })
            else:
                pairs = [f"{c}: {v}" for c, v in zip(col_names, row) if v is not None and str(v).strip()]
                if not pairs:
                    continue
                content = f"[{table}] " + " | ".join(pairs)
                results.append({
                    "content": content,
                    "metadata": {"filename": meta_filename, "table": table},
                })

        logger.info("  ⟹ DuckDB: %d résultat(s) [table=%s, col=%s, filters=%s]",
                    len(results), table, sql_col or "ALL", filters or "-")
        return results

    def _list_values_fallback(self, table, column, filters, distinct, hide_technical=True):
        sql_table = self.tables[table]["sql_table"]
        meta_filename = self.tables[table]["filename"]
        cols = self.tables[table]["columns"]
        user_cols = (self.tables[table].get("user_columns") or cols) if hide_technical else cols
        rows = self.conn.execute(f'SELECT * FROM "{sql_table}"').fetchall()
        sql_col = self._resolve_column(table, column) if column else None

        resolved_filters = []
        for fkey, fval in (filters or {}).items():
            rc = self._resolve_column(table, fkey)
            if rc and fval:
                resolved_filters.append((rc, self._fold(fval)))

        results = []
        seen_distinct: set = set()
        for row in rows:
            kv = dict(zip(cols, row))
            ok = True
            for fc, fv in resolved_filters:
                actual = kv.get(fc)
                if actual is None or fv not in self._fold(str(actual)):
                    ok = False
                    break
            if not ok:
                continue
            if sql_col:
                val = kv.get(sql_col)
                if val is None or not str(val).strip():
                    continue
                key = self._fold(val)
                if distinct and key in seen_distinct:
                    continue
                seen_distinct.add(key)
                results.append({"content": str(val).strip(),
                                "metadata": {"filename": meta_filename, "table": table, "column": sql_col}})
            else:
                pairs = [f"{c}: {v}" for c, v in kv.items()
                         if c in user_cols and v is not None and str(v).strip()]
                if pairs:
                    results.append({"content": f"[{table}] " + " | ".join(pairs),
                                    "metadata": {"filename": meta_filename, "table": table}})
        return results

    def count_rows(self, table: str, filters: Optional[Dict[str, str]] = None) -> int:
        rows = self.list_values(table, column=None, filters=filters, distinct=False)
        return len(rows)

    def reload(self, docs_dir: Optional[str] = None) -> int:
        target_dir = docs_dir or self.docs_dir
        self.conn.close()
        import duckdb
        self.conn = duckdb.connect(":memory:")
        self.tables.clear()
        self.last_warnings = []
        self._load_all(target_dir)
        logger.info("StructuredQueryEngine.reload(): %d table(s) — %s",
                    len(self.tables), ", ".join(self.tables.keys()) or "—")
        return len(self.tables)

    def get_primary_column(self, table: str) -> Optional[str]:
        """Détecte la colonne principale (identifiant/nom) via statistiques pures.

        Critères (sans mots-clés hardcodés) :
        - Ratio d'unicité élevé × longueur moyenne × log(distinct)
        - Exclut les colonnes dont >70 % des valeurs font ≤5 caractères (codes courts)
        - Bonus x1.4 si le nom de la colonne contient le stem de la table
        """
        import math

        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return None

        sql_table  = self.tables[table]["sql_table"]
        user_cols  = self.tables[table].get("user_columns") or self.tables[table]["columns"]

        if not user_cols:
            return None
        if len(user_cols) == 1:
            return user_cols[0]

        table_stem = self._sql_ident(table).replace("_", "").upper()

        best_col:   Optional[str] = None
        best_score: float = -1.0

        for col in user_cols:
            try:
                row = self.conn.execute(
                    f'SELECT COUNT(DISTINCT "{col}"), COUNT("{col}") '
                    f'FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\''
                ).fetchone()
                distinct_count, non_null = (row[0], row[1]) if row else (0, 0)
                if distinct_count == 0:
                    continue

                sample = self.conn.execute(
                    f'SELECT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' LIMIT 30'
                ).fetchall()
                vals = [str(r[0]).strip() for r in sample if r[0]]
                if not vals:
                    continue

                avg_len = sum(len(v) for v in vals) / len(vals)
                if avg_len < 4:
                    continue

                # Ignore columns whose values are mostly short codes (≤5 chars)
                short_ratio = sum(1 for v in vals if len(v) <= 5) / len(vals)
                if short_ratio > 0.7:
                    continue

                ratio_unique = distinct_count / max(non_null, 1)

                # Penalise very long values (job titles, descriptions) — names are
                # typically short (4–15 chars). Quadratic penalty beyond 15 chars
                # ensures NOM/PRENOM columns beat FONCTION/POSTE columns.
                effective_len = avg_len if avg_len <= 15 else 15.0 * (15.0 / avg_len) ** 2
                score = ratio_unique * effective_len * math.log1p(distinct_count)

                col_stem = col.upper().replace("_", "")

                # Penaliser les colonnes dont les valeurs commencent par le stem
                # de la table (ex: "DIRECTION JURIDIQUE" dans table DIRECTION,
                # "SERVICE HELP DESK" dans table SERVICE). Ce sont des colonnes
                # entité/intitulé, pas des colonnes nom de personne.
                if table_stem and len(table_stem) >= 4:
                    try:
                        stem_count = self.conn.execute(
                            f'SELECT COUNT(*) FROM "{sql_table}" '
                            f'WHERE strip_accents(UPPER(CAST("{col}" AS VARCHAR))) '
                            f'LIKE strip_accents(?)',
                            [f"{table_stem} %"],
                        ).fetchone()[0]
                        if non_null > 0 and stem_count / non_null >= 0.3:
                            score *= 0.15  # colonne entité → forte pénalité
                    except Exception:
                        pass

                # Keyword boost — colonnes typiques d'un nom de personne.
                # Fonctionne pour tout nouveau fichier avec un nommage standard.
                _PERSON_KW = {"NOM", "PRENOM", "AGENT", "EMPLOYE", "TITULAIRE",
                              "GESTIONNAIRE", "COLLABORATEUR", "PERSONNEL",
                              "RESPONSABLE"}
                if any(kw in col_stem for kw in _PERSON_KW):
                    score *= 2.0

                if score > best_score:
                    best_score = score
                    best_col   = col

            except Exception as e:
                logger.debug("  get_primary_column: erreur col %s: %s", col, e)
                continue

        if best_col is None and user_cols:
            best_col = user_cols[0]

        logger.info("get_primary_column(%s) → %s", table, best_col)
        return best_col

    def get_role_column(self, table: str) -> Optional[str]:
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return None

        sql_table   = self.tables[table]["sql_table"]
        user_cols   = self.tables[table].get("user_columns") or self.tables[table]["columns"]
        row_count   = self.tables[table]["row_count"]
        primary_col = self.get_primary_column(table)

        if not user_cols or row_count == 0:
            return None

        best_col:   Optional[str] = None
        best_score: float = -1.0

        for col in user_cols:
            if col == primary_col:
                continue
            try:
                row = self.conn.execute(
                    f'SELECT COUNT(DISTINCT "{col}"), COUNT("{col}") '
                    f'FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\''
                ).fetchone()
                distinct_count, non_null = (row[0], row[1]) if row else (0, 0)
                if distinct_count < 2 or non_null == 0:
                    continue

                sample = self.conn.execute(
                    f'SELECT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' LIMIT 10'
                ).fetchall()
                vals    = [str(r[0]).strip() for r in sample if r[0]]
                avg_len = sum(len(v) for v in vals) / max(len(vals), 1)
                if avg_len < 4:
                    continue

                cardinality_ratio = distinct_count / max(non_null, 1)
                # Skip: high cardinality + short values → IDs, codes, short names.
                # Allow: high cardinality + long values → unique job titles/functions.
                if cardinality_ratio > 0.5 and avg_len < 12:
                    continue

                # Low-cardinality repeating roles get a small cardinality bonus.
                # Unique long titles (CHEF DE DEPARTEMENT…) are slightly penalised
                # but still score higher than short status fields (CONFIRME, OUI…).
                if cardinality_ratio <= 0.5:
                    score = avg_len * (1.0 - cardinality_ratio * 0.5)
                else:
                    score = avg_len * 0.4

                # Keyword boost — colonnes typiques d'un rôle/fonction.
                _ROLE_KW = {"FONCTION", "POSTE", "GRADE", "TITRE", "QUALIFICATION",
                            "EMPLOI", "MISSION", "CATEGORIE", "SPECIALITE", "METIER"}
                col_up = col.upper().replace("_", "")
                if any(kw in col_up for kw in _ROLE_KW):
                    score *= 2.5

                if score > best_score:
                    best_score = score
                    best_col   = col

            except Exception:
                continue

        logger.info("get_role_column(%s) → %s", table, best_col)
        return best_col

    def find_column_for_question(
        self,
        table: str,
        question: str,
        excluded_cols: Optional[List[str]] = None,
    ) -> Optional[str]:
        """Trouve dynamiquement la colonne que la question cible.

        Entièrement sans vocabulaire hardcodé — fonctionne pour tout fichier,
        toute langue, tout nommage de colonnes.

        Étape 1 — Direct name match (score 10) :
          Le nom de colonne replié (≥ 4 chars) apparaît verbatim dans la question.

        Étape 2 — Token prefix match (score proportionnel) :
          Chaque token de la question est comparé aux tokens du nom de colonne.
          Si l'un est préfixe de l'autre (min 3 chars), score += longueur commune /
          longueur max. Ex : "salaire" ↔ "SAL_NET" → "salaire".startswith("sal") ✓

        Étape 3 — Sample value match (fallback) :
          Activé seulement si les étapes 1+2 donnent score=0.
          Les tokens de la question sont recherchés dans les valeurs-exemples
          stockées à la phase de chargement.
        """
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return None

        columns   = self.tables[table]["columns"]
        excluded  = set(excluded_cols or [])
        candidates = [c for c in columns if c not in excluded]
        if not candidates:
            return None

        q_folded = self._fold(question)
        scores: Dict[str, float] = {c: 0.0 for c in candidates}

        # ── Étape 1 : nom de colonne présent dans la question ─────────────
        for col in candidates:
            cf  = self._fold(col.replace("_", " "))
            cf2 = self._fold(col.replace("_", ""))
            if len(cf2) >= 4 and (cf in q_folded or cf2 in q_folded):
                scores[col] += 10.0

        # ── Étape 2 : correspondance de préfixes de tokens ─────────────────
        q_tokens = [
            t for t in re.split(r"[_\s\-/,;:!?.()]+", q_folded)
            if len(t) >= 3 and t not in self._MATCH_STOP
        ]
        for col in candidates:
            col_tokens = [
                t for t in re.split(r"[_\s\-/,;:!?.()]+", self._fold(col))
                if len(t) >= 3
            ]
            for qt in q_tokens:
                for ct in col_tokens:
                    min_len = min(len(qt), len(ct))
                    if min_len < 3:
                        continue
                    if qt.startswith(ct) or ct.startswith(qt):
                        prefix_len = sum(1 for a, b in zip(qt, ct) if a == b)
                        scores[col] += prefix_len / max(len(qt), len(ct))

        # ── Étape 3 : correspondance dans les valeurs-exemples (fallback) ──
        if max(scores.values(), default=0.0) == 0.0:
            col_samples = self.tables[table].get("column_samples", {})
            for col in candidates:
                for val in col_samples.get(col, []):
                    vf = self._fold(val)
                    if any(qt in vf for qt in q_tokens if len(qt) >= 3):
                        scores[col] += 1.0

        best = max(scores, key=lambda c: scores[c], default=None)
        if best and scores[best] > 0.0:
            logger.debug("find_column_for_question(%s, %r) → %s (score=%.2f)",
                         table, question[:50], best, scores[best])
            return best
        return None

    def keyword_search(
        self,
        table: str,
        question: str,
        max_results: int = 5,
    ) -> List[Dict[str, Any]]:
        """Recherche par mots-clés dans une table DuckDB.

        Stratégie en deux phases :
        1. AND : tous les tokens doivent matcher → précision maximale.
        2. OR pondéré : fallback si AND échoue ; chaque ligne est scorée par
           la rareté de ses tokens correspondants (moins fréquent = plus spécifique).
        """
        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return []

        sql_table     = self.tables[table]["sql_table"]
        meta_filename = self.tables[table]["filename"]
        user_cols     = self.tables[table].get("user_columns") or self.tables[table]["columns"]
        all_cols      = self.tables[table]["columns"]
        user_cols_set = set(user_cols)

        # Mots vides grammaticaux uniquement — aucun vocabulaire métier hardcodé.
        _STOP = {
            # Interrogatifs / relatifs
            "qui", "quel", "quelle", "quels", "quelles", "comment",
            "pourquoi", "combien", "quand", "lequel", "laquelle", "que", "qu",
            # Démonstratifs
            "ce", "cet", "cette", "ces", "ceci", "cela",
            # Articles / déterminants
            "le", "la", "les", "un", "une", "des", "du",
            "mon", "ton", "son", "ma", "ta", "sa", "mes", "tes", "ses",
            "nos", "vos", "leur", "leurs",
            # Contractions / élisions
            "d", "l", "j", "m", "n", "s", "t", "y", "c",
            # Prépositions
            "de", "dans", "par", "pour", "sur", "avec", "sans",
            "en", "au", "aux", "vers", "entre", "parmi", "selon", "sous",
            "dont", "lors", "dès", "des", "chez", "hors",
            # Conjonctions / adverbes
            "et", "ou", "mais", "donc", "car", "si", "or", "ni",
            "bien", "aussi", "tres", "plus", "moins", "tout", "tous",
            # Verbes d'état / courants
            "est", "sont", "etait", "etaient", "avoir", "etre", "fait",
            "peut", "doit", "faut", "a", "ont", "va", "vas",
            # Anglais fonctionnel
            "the", "and", "for", "with", "from", "who", "what", "is", "are",
        }

        # Tokenisation : supprime la ponctuation avant le filtrage
        raw_tokens = re.split(r"[\s\-_/,;:!?.]+", self._fold(question))
        tokens = [t for t in raw_tokens if len(t) >= 3 and t not in _STOP]

        # Filtre les tokens qui sont le nom de la table ou ses formes plurielles
        # (ex: "services" dans la question vers la table SERVICE) — ils matcheraient
        # toutes les lignes et n'apportent aucune discrimination.
        table_toks = {
            t for t in re.split(r"[\s_]+", self._fold(table.replace("_", " ")))
            if len(t) >= 3
        }
        tokens = [
            t for t in tokens
            if t not in table_toks
            and not any(t.startswith(tt) and len(t) <= len(tt) + 2 for tt in table_toks)
            and not any(tt.startswith(t) and len(t) >= 4 for tt in table_toks)
        ]

        if not tokens:
            return []

        # SELECT toutes les colonnes → raw_row complet (attributs techniques inclus).
        # La recherche WHERE utilise user_cols uniquement (évite les faux positifs sur IDs).
        select_clause = ", ".join(f'"{c}"' for c in all_cols)

        # ── Phase 1 : AND — tous les tokens doivent matcher dans la ligne ──
        and_conds: List[str] = []
        and_params: List[str] = []
        for token in tokens:
            col_or = " OR ".join(
                f"{self._like_col(col)} LIKE {self._LIKE_PARAM}"
                for col in user_cols
            )
            and_conds.append(f"({col_or})")
            and_params.extend([f"%{token}%"] * len(user_cols))

        _and_matched = False
        try:
            rows = self.conn.execute(
                f'SELECT {select_clause} FROM "{sql_table}" '
                f'WHERE {" AND ".join(and_conds)} LIMIT {max_results * 4}',
                and_params,
            ).fetchall()
            _and_matched = len(rows) > 0
        except Exception as e:
            logger.warning("keyword_search AND échoué (%s), fallback OR+score", e)
            rows = []

        # ── Phase 1b : tri par précision sur la colonne rôle ─────────────────
        # On calcule deux sous-scores :
        #   1. prefix_bonus : +2 si la valeur du rôle COMMENCE par les tokens
        #      de la question (ex: FONCTION="CHEF DE DEPARTEMENT..." pour
        #      "Chef du Département HSE"). Évite que "ingenieur en CHEF" batte
        #      "CHEF DE DEPARTEMENT" alors que la question demande un "chef de département".
        #   2. matched - extra*0.5 : score standard token overlap.
        # Le sous-score 1 prime (poids x3) pour lever les ambiguïtés de titres.
        if _and_matched and len(rows) > 1:
            q_token_set   = set(tokens)
            role_col_sort = self.get_role_column(table)  # None si non détecté

            def _precision(row):
                rv = {c: str(v).strip() if v is not None else ""
                      for c, v in zip(all_cols, row)}
                if role_col_sort and rv.get(role_col_sort):
                    score_text = self._fold(rv[role_col_sort])
                else:
                    score_text = self._fold(
                        " ".join(rv[c] for c in user_cols if rv.get(c))
                    )
                row_toks = {
                    t for t in re.split(r"[\s\-_/,;:!?.]+", score_text)
                    if len(t) >= 3 and t not in _STOP
                }
                matched = len(row_toks & q_token_set)
                extra   = len(row_toks - q_token_set)
                base    = matched - extra * 0.5

                # Bonus de préfixe : la FONCTION commence-t-elle par les premiers
                # tokens de la question (dans le même ordre) ?
                # Ex: q_tokens=["chef","departement","hse"] et
                #     role="chef de departement hse" → préfixe OK → bonus +3
                prefix_bonus = 0.0
                if role_col_sort and rv.get(role_col_sort):
                    role_f = self._fold(rv[role_col_sort])
                    role_start_toks = [
                        t for t in re.split(r"[\s\-_/,;:!?.]+", role_f)
                        if len(t) >= 3
                    ]
                    # Compte combien de tokens question matchent en début de rôle
                    common_prefix = 0
                    for qt in q_token_set:
                        if role_start_toks and role_start_toks[0].startswith(qt):
                            common_prefix += 1
                        elif len(role_start_toks) > 1 and role_start_toks[1].startswith(qt):
                            common_prefix += 0.5
                    prefix_bonus = common_prefix * 3.0

                return base + prefix_bonus

            rows = sorted(rows, key=_precision, reverse=True)[:max_results]

        # ── Phase 2 : OR pondéré par rareté si AND ne donne rien ──────────
        if not rows:
            or_parts: List[str] = []
            or_params: List[str] = []
            for col in user_cols:
                for token in tokens:
                    or_parts.append(
                        f"{self._like_col(col)} LIKE {self._LIKE_PARAM}"
                    )
                    or_params.append(f"%{token}%")

            try:
                candidates = self.conn.execute(
                    f'SELECT {select_clause} FROM "{sql_table}" '
                    f'WHERE {" OR ".join(or_parts)} LIMIT {max_results * 6}',
                    or_params,
                ).fetchall()
            except Exception as e:
                logger.warning("keyword_search OR échoué (%s)", e)
                return []

            # Poids = rareté du token (1 / nb_lignes_correspondantes)
            # Plus le token est rare → plus il est discriminant.
            token_weights: Dict[str, float] = {}
            for token in tokens:
                col_or = " OR ".join(
                    f"{self._like_col(col)} LIKE {self._LIKE_PARAM}"
                    for col in user_cols
                )
                try:
                    count = self.conn.execute(
                        f'SELECT COUNT(*) FROM "{sql_table}" WHERE {col_or}',
                        [f"%{token}%"] * len(user_cols),
                    ).fetchone()[0]
                except Exception:
                    count = 1
                token_weights[token] = 1.0 / max(count, 1)

            scored: List[tuple] = []
            for row in candidates:
                raw_row = {c: str(v).strip() if v is not None else ""
                           for c, v in zip(all_cols, row)}
                # Score sur les colonnes visibles uniquement (évite bruit sur IDs)
                user_text = self._fold(" ".join(raw_row[c] for c in user_cols if raw_row.get(c)))
                score = sum(w for t, w in token_weights.items() if t in user_text)
                scored.append((score, row, raw_row))

            scored.sort(key=lambda x: -x[0])
            rows = [s[1] for s in scored[:max_results]]

        results: List[Dict[str, Any]] = []
        for row in rows:
            # raw_row contient TOUTES les colonnes — attributs techniques inclus
            # (ex : SALAIRE numérique) pour que _find_target_column puisse les lire.
            raw_row = {c: str(v).strip() if v is not None else "" for c, v in zip(all_cols, row)}
            # content n'affiche que les colonnes visibles par l'utilisateur
            pairs   = [f"{c}: {v}" for c, v in raw_row.items() if c in user_cols_set and v]
            if not pairs:
                continue
            results.append({
                "content":  f"[{table}] " + " | ".join(pairs),
                "metadata": {
                    "filename":  meta_filename,
                    "table":     table,
                    "raw_row":   raw_row,
                    "and_match": _and_matched,
                },
            })

        logger.info("keyword_search(%s, %r) → %d résultat(s)", table, tokens, len(results))
        return results

    def get_entity_column(self, table: str) -> Optional[str]:
        """Colonne la plus représentative pour lister les entités d'une table.

        Contrairement à get_primary_column() qui préfère les valeurs courtes
        (noms de personnes), cette méthode préfère les valeurs longues et uniques
        (intitulés, désignations, noms d'entités).
        Utilisée exclusivement par le chemin exhaustif (listes).
        """
        import math

        table = self._normalize_stem(table) if table not in self.tables else table
        if table not in self.tables:
            return None

        sql_table  = self.tables[table]["sql_table"]
        user_cols  = self.tables[table].get("user_columns") or self.tables[table]["columns"]

        if not user_cols:
            return None
        if len(user_cols) == 1:
            return user_cols[0]

        table_stem = self._sql_ident(table).replace("_", "").upper()

        # Heuristique primaire : colonne dont ≥ 40 % des valeurs commencent
        # par le stem de la table suivi d'un espace ("SERVICE ", "DIRECTION "…).
        # Fiable même quand le nom de colonne est générique (INTITULE, LIBELLE…)
        # et immunisée contre les colonnes CHANTIER / LIEU qui ont des valeurs longues.
        if len(table_stem) >= 5:
            stem_prefix = table_stem + " "
            stem_best_col: Optional[str] = None
            stem_best_ratio: float = 0.0
            for col in user_cols:
                try:
                    sample = self.conn.execute(
                        f'SELECT "{col}" FROM "{sql_table}" '
                        f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' LIMIT 60'
                    ).fetchall()
                    vals = [str(r[0]).strip().upper() for r in sample if r[0]]
                    if not vals:
                        continue
                    ratio = sum(1 for v in vals if v.startswith(stem_prefix)) / len(vals)
                    if ratio > stem_best_ratio:
                        stem_best_ratio = ratio
                        stem_best_col = col
                except Exception:
                    continue
            if stem_best_col and stem_best_ratio >= 0.4:
                logger.info("get_entity_column(%s) → %s (stem-prefix %.0f%%)",
                            table, stem_best_col, stem_best_ratio * 100)
                return stem_best_col

        best_col:   Optional[str] = None
        best_score: float = -1.0

        for col in user_cols:
            try:
                row = self.conn.execute(
                    f'SELECT COUNT(DISTINCT "{col}"), COUNT("{col}") '
                    f'FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\''
                ).fetchone()
                distinct_count, non_null = (row[0], row[1]) if row else (0, 0)
                if distinct_count < 2:
                    continue

                sample = self.conn.execute(
                    f'SELECT "{col}" FROM "{sql_table}" '
                    f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") <> \'\' LIMIT 30'
                ).fetchall()
                vals = [str(r[0]).strip() for r in sample if r[0]]
                if not vals:
                    continue

                avg_len = sum(len(v) for v in vals) / len(vals)
                if avg_len < 5:
                    continue

                # Ignore les colonnes à valeurs courtes (codes, abréviations ≤ 6 chars)
                short_ratio = sum(1 for v in vals if len(v) <= 6) / len(vals)
                if short_ratio > 0.6:
                    continue

                ratio_unique = distinct_count / max(non_null, 1)
                # Formule originale sans pénalité longueur → favorise les longs intitulés
                score = ratio_unique * avg_len * math.log1p(distinct_count)

                col_stem = col.upper().replace("_", "")

                # Boost fort si le nom de colonne contient le stem de la table
                if table_stem and len(table_stem) >= 5 and \
                        (table_stem in col_stem or col_stem in table_stem):
                    score *= 3.0

                # Keyword boost — colonnes typiques d'un intitulé/désignation
                # d'entité. Rend la détection robuste pour tout nouveau fichier.
                _ENTITY_KW = {"INTITULE", "LIBELLE", "DESIGNATION", "DENOMINATION",
                              "OBJET", "DESCRIPTION", "THÈME", "THEME", "SUJET",
                              "TITRE", "REFERENCE", "FORMATION", "ACTIVITE"}
                if any(kw in col_stem for kw in _ENTITY_KW):
                    score *= 2.5

                if score > best_score:
                    best_score = score
                    best_col   = col

            except Exception:
                continue

        if best_col is None and user_cols:
            best_col = user_cols[0]

        logger.info("get_entity_column(%s) → %s", table, best_col)
        return best_col

    def close(self) -> None:
        try:
            self.conn.close()
        except Exception:
            pass