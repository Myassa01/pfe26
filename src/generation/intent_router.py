"""Routage intelligent des requêtes : découverte de schéma + classification LLM."""

import json
import logging
import re
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from .llm import HFClient
from ..ingestion.loader import _detect_header_row

logger = logging.getLogger(__name__)


class SchemaDiscovery:
    EXCEL_EXTS = {".xlsx", ".xls"}
    DOC_EXTS   = {".docx", ".doc", ".pdf", ".txt", ".md", ".html", ".htm"}

    def __init__(self, docs_dir: str, max_samples: int = 5, max_sample_len: int = 40):
        self.docs_dir       = docs_dir
        self.max_samples    = max_samples
        self.max_sample_len = max_sample_len

    def scan(self) -> Dict[str, dict]:
        schema: Dict[str, dict] = {}
        if not Path(self.docs_dir).exists():
            logger.warning("SchemaDiscovery: docs_dir introuvable: %s", self.docs_dir)
            return schema
        for file in sorted(Path(self.docs_dir).rglob("*")):
            if not file.is_file():
                continue
            ext  = file.suffix.lower()
            stem = self._normalize_stem(file.name)
            if ext in self.EXCEL_EXTS:
                entry = self._scan_excel(file)
                if entry:
                    schema[stem] = entry
            elif ext in self.DOC_EXTS:
                schema[stem] = {
                    "columns": [], "samples": {}, "is_doc": True,
                    "filename": file.name,
                }
        logger.info("SchemaDiscovery: %d sources détectées (%s)",
                    len(schema), ", ".join(schema.keys()))
        return schema

    @staticmethod
    def _normalize_stem(fname: str) -> str:
        stem = fname.rsplit(".", 1)[0] if "." in fname else fname
        stem = stem.upper().strip()
        stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
        stem = re.sub(r"\s*_\d+\s*$",    "", stem)
        return stem.strip()

    def _scan_excel(self, path: Path) -> Optional[dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
<<<<<<< HEAD
            ws = wb.active or wb.worksheets[0]
=======
            ws = wb.active
>>>>>>> 523536e19cd5c29d340be65ba01ccf0c173c0000
            header_row = _detect_header_row(ws)
            headers: List[str] = [
                str(cell.value).strip() if cell.value else ""
                for cell in ws[header_row]
            ]
            headers = [h for h in headers if h]
            samples: Dict[str, List[str]] = {h: [] for h in headers}
            for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True)):
                if i >= 200:
                    break
                for h, v in zip(headers, row):
                    if v is None or len(samples[h]) >= self.max_samples:
                        continue
                    s = str(v).strip()[:self.max_sample_len]
                    if s and s not in samples[h]:
                        samples[h].append(s)
            wb.close()
            return {
                "columns": headers,
                "samples": {h: samples[h] for h in headers if samples[h]},
                "is_doc":  False,
                "filename": path.name,
            }
        except Exception as e:
            logger.warning("SchemaDiscovery: échec scan %s: %s", path.name, e)
            return None


# ── PROMPTS ────────────────────────────────────────────────────────────────────

_SYSTEM_PROMPT = (
    "Tu es un classifieur de requêtes en français. "
    "Tu retournes UNIQUEMENT un objet JSON valide, rien d'autre."
)

_PROMPT_TEMPLATE = """\
Classifie la question utilisateur en JSON selon le schéma ci-dessous.

Sources disponibles :
{schema_block}

Format JSON attendu (strict, une seule ligne) :
{{"intent":"list|detail|qa","source":"<NOM_SOURCE_OU_null>","column":"<NOM_COLONNE_OU_null>","exhaustive":true|false,"filter":null}}

Règles STRICTES :
- "intent":"list"   → énumération ("liste", "tous", "donne-moi les", "quels sont", "combien").
<<<<<<< HEAD
- "intent":"detail" → explication ou procédure pas à pas ("explique", "détails", "décris", "comment", "comment faire", "procédure", "étapes", "démarche").
=======
- "intent":"detail" → explication ("explique", "détails", "décris").
>>>>>>> 523536e19cd5c29d340be65ba01ccf0c173c0000
- "intent":"qa"     → question ciblée ("qui est", "quel est", "quelle est", "qui sont").
- "source"  : EXACTEMENT un nom de table parmi la liste ci-dessus, ou null.
- "column"  : null pour "qa" et "list" (on retourne toute la ligne).
- "exhaustive" : true UNIQUEMENT si intent="list" ET source est une table Excel. false sinon.
- "filter"  : TOUJOURS null pour intent="qa" — le pipeline cherchera dans toutes les lignes.
  Pour intent="list" avec contrainte : {{"NOM_COLONNE_REEL":"valeur"}}.
  Utilise UNIQUEMENT des noms de colonnes présents dans le schéma.

IMPORTANT — intent="qa" :
  Ne jamais mettre exhaustive=true pour une question "qui est X" ou "quel est X".
  Ne jamais inventer un filtre pour intent="qa" — mettre filter=null.
  Le pipeline se charge de trouver la bonne ligne dans la table.

Exemples :
{examples_block}

Question : {question}
JSON :"""


class IntentRouter:
    def __init__(self, llm: HFClient, schema: Dict[str, dict], cache_size: int = 256):
        self.llm        = llm
        self.schema     = schema
        self._schema_block   = self._build_schema_block(schema)
        self._examples_block = self._build_dynamic_examples(schema)
        self._cache: OrderedDict = OrderedDict()
        self._cache_size = cache_size
        logger.info("IntentRouter: %d source(s) dans le schéma (%s)",
                    len(schema), ", ".join(schema.keys()))

    def classify(self, question: str) -> dict:
        key = self._normalize_question(question)
        if key in self._cache:
            self._cache.move_to_end(key)
            logger.info("IntentRouter: cache hit pour %r", question[:60])
            return self._cache[key]

        prompt = _PROMPT_TEMPLATE.format(
            schema_block=self._schema_block,
            examples_block=self._examples_block,
            question=question.strip(),
        )
        try:
            raw = self.llm.generate(
                prompt=prompt,
                system=_SYSTEM_PROMPT,
                temperature=0.0,
                max_tokens=150,
            )
        except Exception as e:
            logger.warning("IntentRouter: appel LLM échoué (%s) — fallback qa", e)
            return self._fallback()

        parsed = self._parse_json(raw)
        result = self._validate(parsed) if parsed else self._fallback()
        logger.info(
            "IntentRouter: intent=%s source=%s column=%s exhaustive=%s filter=%s",
            result["intent"], result["source"], result["column"],
            result["exhaustive"], result["filter"],
        )

        self._cache[key] = result
        if len(self._cache) > self._cache_size:
            self._cache.popitem(last=False)

        return result

    def _build_schema_block(self, schema: Dict[str, dict]) -> str:
        if not schema:
            return "(aucune source disponible)"
        lines = []
        for name, info in schema.items():
            if info.get("is_doc"):
                lines.append(f"* {name} (document texte) — descriptions, explications")
                continue
            row_count = info.get("row_count", "?")
            cols      = info.get("columns", [])
            samples   = info.get("samples",  {})
            lines.append(f"* {name} ({row_count} lignes) :")
            for col in cols:
                vals = samples.get(col, [])
                if vals:
                    sample_str = ", ".join(f'"{v}"' for v in vals[:5])
                    lines.append(f"    - {col} → ex: {sample_str}")
                else:
                    lines.append(f"    - {col}")
        return "\n".join(lines)

    def _build_dynamic_examples(self, schema: Dict[str, dict]) -> str:
        if not schema:
            return "(aucun exemple disponible)"
        lines: List[str] = []

        for table_name, info in schema.items():
            if info.get("is_doc"):
                lines.append(
                    f'Q: "Explique {table_name.lower()}"\n'
                    f'JSON: {{"intent":"detail","source":"{table_name}",'
                    f'"column":null,"exhaustive":false,"filter":null}}'
                )
                continue

            cols    = info.get("columns", [])
            samples = info.get("samples", {})
            if not cols:
                continue

            # ── Exemple 1 : liste exhaustive ──────────────────────────────
            lines.append(
                f'Q: "Liste tous les {table_name.lower()}"\n'
                f'JSON: {{"intent":"list","source":"{table_name}",'
                f'"column":null,"exhaustive":true,"filter":null}}'
            )

            # ── Exemple 2 : question QA "qui est" — filter=null toujours ──
            # On prend la première colonne avec des valeurs texte
            name_col = self._pick_target_column(cols, samples, exclude=None)
            if name_col:
                sample_vals = samples.get(name_col, [])
                sample_val  = sample_vals[0] if sample_vals else "X"
                lines.append(
                    f'Q: "Qui est le responsable de {sample_val} ?"\n'
                    f'JSON: {{"intent":"qa","source":"{table_name}",'
                    f'"column":null,"exhaustive":false,"filter":null}}'
                )
                lines.append(
                    f'Q: "Quel est le {name_col.lower()} de {sample_val} ?"\n'
                    f'JSON: {{"intent":"qa","source":"{table_name}",'
                    f'"column":null,"exhaustive":false,"filter":null}}'
                )

            # ── Exemple 3 : liste filtrée ─────────────────────────────────
            filter_col, filter_val = self._pick_filter_column(cols, samples, table_name)
            if filter_col and filter_val:
                lines.append(
                    f'Q: "{table_name.lower()} du {filter_val}"\n'
                    f'JSON: {{"intent":"list","source":"{table_name}",'
                    f'"column":null,"exhaustive":true,'
                    f'"filter":{{"{filter_col}":"{filter_val}"}}}}'
                )

            lines.append("")

        return "\n".join(lines).strip()

    def _pick_filter_column(
        self,
        cols: List[str],
        samples: Dict[str, List[str]],
        table_name: str,
    ) -> Tuple[Optional[str], Optional[str]]:
        best_col: Optional[str] = None
        best_val: Optional[str] = None
        best_score: float = -1.0

        for col in cols:
            vals = [v for v in samples.get(col, []) if v and len(v) >= 2]
            if not vals:
                continue
            if all(v.replace(".", "").replace(",", "").replace("-", "").isdigit()
                   for v in vals):
                continue
            avg_len = sum(len(v) for v in vals) / len(vals)
            if avg_len > 40 or avg_len < 2:
                continue
            col_up = col.upper()
            bonus = 1.0
            for kw in ("DIRECTION", "DEPARTEMENT", "SERVICE", "TYPE", "CATEGORIE",
                       "STATUT", "STATUS", "ACTIVITE", "ZONE", "REGION", "SITE",
                       "WILAYA", "UNITE", "GROUPE"):
                if kw in col_up:
                    bonus = 2.0
                    break
            table_stem = table_name.upper().replace("_", "")
            col_stem   = col_up.replace("_", "")
            if table_stem in col_stem or col_stem in table_stem:
                bonus *= 0.5
            score = bonus / avg_len
            if score > best_score:
                best_score = score
                best_col   = col
                best_val   = vals[0]

        return best_col, best_val

    @staticmethod
    def _pick_target_column(
        cols: List[str],
        samples: Dict[str, List[str]],
        exclude: Optional[str] = None,
    ) -> Optional[str]:
        NAME_KWS = ("NOM", "LIBELLE", "PRENOM", "INTITULE", "DESIGNATION",
                    "TITRE", "LABEL", "CHANTIER")
        for col in cols:
            if col == exclude:
                continue
            col_up = col.upper()
            for kw in NAME_KWS:
                if kw in col_up:
                    return col
        for col in cols:
            if col == exclude:
                continue
            vals = samples.get(col, [])
            if vals and not all(
                v.replace(".", "").replace(",", "").replace("-", "").isdigit()
                for v in vals
            ):
                return col
        return None

    @staticmethod
    def _normalize_question(q: str) -> str:
        q = q.lower().strip()
        for src, dst in [
            ("é","e"),("è","e"),("ê","e"),("ë","e"),
            ("à","a"),("â","a"),("ä","a"),
            ("î","i"),("ï","i"),("ô","o"),("ö","o"),
            ("ù","u"),("û","u"),("ü","u"),("ç","c"),
        ]:
            q = q.replace(src, dst)
        return re.sub(r"\s+", " ", q)

    def _parse_json(self, raw: str) -> Optional[dict]:
        if not raw:
            return None
        match     = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        candidate = match.group(0) if match else raw.strip()
        for attempt in (candidate, self._repair_json(candidate)):
            try:
                return json.loads(attempt)
            except (json.JSONDecodeError, TypeError):
                continue
        return None

    @staticmethod
    def _repair_json(s: str) -> str:
        s = re.sub(r"'", '"', s)
        s = re.sub(r",\s*([}\]])", r"\1", s)
        return s

    def _validate(self, data: dict) -> dict:
        # ── intent ────────────────────────────────────────────────────────
        intent = str(data.get("intent", "qa")).lower()
        if intent not in ("list", "detail", "qa"):
            intent = "qa"

        # ── source ────────────────────────────────────────────────────────
        source = data.get("source")
        if source in (None, "null", "None", ""):
            source = None
        else:
            source = self._resolve_source(str(source))

        # ── column ────────────────────────────────────────────────────────
        column = data.get("column")
        if column in (None, "null", "None", ""):
            column = None
        elif source and not self.schema[source].get("is_doc"):
            column = self._resolve_column(source, str(column))

        # ── exhaustive — UNIQUEMENT pour tables structurées ───────────────
        is_structured = (
            source is not None
            and not self.schema.get(source, {}).get("is_doc", True)
        )
        # qa ne doit JAMAIS être exhaustive
        exhaustive = (
            (bool(data.get("exhaustive", False)) or intent == "list")
            and intent != "qa"
            and is_structured
        )

        # ── filter — jamais pour qa ────────────────────────────────────────
        filt = data.get("filter")
        if intent == "qa":
            filt = None  # forcer null pour qa — le pipeline fait full-scan
        elif not isinstance(filt, dict) or not filt:
            filt = None
        elif source and not self.schema[source].get("is_doc"):
            cols_up = {
                c.upper(): c
                for c in self.schema[source].get("columns", [])
            }
            resolved_filt: Dict[str, str] = {}
            for k, v in filt.items():
                real_col = self._fuzzy_resolve_column(str(k), cols_up)
                if real_col:
                    resolved_filt[real_col] = str(v)
                else:
                    logger.warning(
                        "IntentRouter: colonne de filtre inconnue '%s' ignorée "
                        "(table %s, colonnes: %s)",
                        k, source, ", ".join(cols_up.keys()),
                    )
            filt = resolved_filt or None

        return {
            "intent":     intent,
            "source":     source,
            "column":     column,
            "exhaustive": exhaustive,
            "filter":     filt,
            "confidence": "high",
        }

    def _resolve_source(self, name: str) -> Optional[str]:
        if name in self.schema:
            return name
        up = name.upper().strip()
        if up in self.schema:
            return up
        for key in self.schema:
            if key in up or up in key:
                return key
        logger.warning("IntentRouter: source '%s' inconnue (schéma: %s)",
                       name, ", ".join(self.schema.keys()))
        return None

    def _resolve_column(self, source: str, col_name: str) -> Optional[str]:
        cols_up = {c.upper(): c for c in self.schema[source].get("columns", [])}
        return self._fuzzy_resolve_column(col_name, cols_up)

    @staticmethod
    def _fuzzy_resolve_column(col_name: str, cols_up: Dict[str, str]) -> Optional[str]:
        target = col_name.upper().strip()
        if target in cols_up:
            return cols_up[target]
        for key, original in cols_up.items():
            if target in key or key in target:
                return original
        for key, original in cols_up.items():
            min_len = min(len(target), len(key))
            if min_len >= 4 and target[:min_len] == key[:min_len]:
                return original
        return None

    @staticmethod
    def _fallback() -> dict:
        return {
            "intent":     "qa",
            "source":     None,
            "column":     None,
            "exhaustive": False,
            "filter":     None,
            "confidence": "low",
        }
